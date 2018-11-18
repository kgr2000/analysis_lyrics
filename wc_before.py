import pandas as pd
from openpyxl import Workbook
wb = Workbook()
from openpyxl.utils.dataframe import dataframe_to_rows

df = pd.read_excel('d:/python/study/lyrics/bts_1.xlsx', encoding='euckr')
ws = wb.create_sheet(title = "bts")

for row in dataframe_to_rows(df, index=True, header=True):
    if len(row) > 1:
        p = round((row[2])/10)
        for num in range(1, p) :
            ws.append(row)

#ws2 = wb.create_sheet(title = "bts_2")
#words = []
#counts = []

#for w in a.index :
#    words.append(w)
#for c in a.values :
#    counts.append(c)

#ws1.append(words)
#ws2.append(counts)

wb.save("D:/python/study/lyrics/bts_analysis_2.xlsx") #저장
